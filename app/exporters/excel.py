"""Bounded and formula-injection-safe Excel export."""

from collections.abc import Sequence
from io import BytesIO
from typing import cast

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.domain.exports import ExportFormat, ExportItem, ExportMetadata, RenderedExport
from app.exporters.common import (
    CATEGORY_LABELS,
    excel_safe_text,
    format_time,
    safe_filename,
    safe_hyperlink,
    safe_sheet_name,
)


class ExcelExporter:
    export_format = ExportFormat.EXCEL
    max_items = 10_000
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    _headers = (
        "序号",
        "标题",
        "最终分类",
        "分类来源",
        "来源名称",
        "发布时间",
        "发现时间",
        "简介",
        "原文链接",
        "收藏状态",
        "自动分类分数",
        "自动分类原因",
    )
    _widths = (8, 44, 20, 12, 24, 19, 19, 60, 14, 12, 16, 48)

    def render(
        self,
        items: Sequence[ExportItem],
        metadata: ExportMetadata,
    ) -> RenderedExport:
        workbook = Workbook()
        buffer = BytesIO()
        try:
            sheet = cast(Worksheet, workbook.active)
            sheet.title = safe_sheet_name("资讯列表")
            sheet.append(self._headers)
            header_fill = PatternFill(fill_type="solid", fgColor="176B68")
            for raw_cell in sheet[1]:
                cell = cast(Cell, raw_cell)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for index, item in enumerate(items, start=1):
                sheet.append(
                    (
                        index,
                        excel_safe_text(item.title),
                        CATEGORY_LABELS[item.effective_category],
                        item.category_origin,
                        excel_safe_text(item.source_name),
                        format_time(item.published_at),
                        format_time(item.discovered_at),
                        excel_safe_text(item.summary or ""),
                        "查看原文" if safe_hyperlink(item.original_url) else "链接不可用",
                        "是" if item.is_favorite else "否",
                        item.classification_score,
                        excel_safe_text(item.classification_reason or ""),
                    )
                )
                row = index + 1
                link = safe_hyperlink(item.original_url)
                if link is not None:
                    link_cell = cast(Cell, sheet.cell(row=row, column=9))
                    link_cell.hyperlink = link
                    link_cell.style = "Hyperlink"
                for raw_cell in sheet[row]:
                    cell = cast(Cell, raw_cell)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = f"A1:L{len(items) + 1}"
            sheet.row_dimensions[1].height = 24
            for index, width in enumerate(self._widths, start=1):
                sheet.column_dimensions[get_column_letter(index)].width = width

            notes = workbook.create_sheet(safe_sheet_name("导出说明"))
            notes.column_dimensions["A"].width = 18
            notes.column_dimensions["B"].width = 90
            notes.append(("项目", "内容"))
            note_rows = (
                ("生成时间", format_time(metadata.generated_at)),
                ("筛选条件", excel_safe_text(metadata.filter_summary)),
                ("导出条数", metadata.item_count),
                ("分类说明", "最终分类优先采用人工分类; 没有人工分类时使用自动分类。"),
                ("使用提示", "数据仅供内部参考, 请通过原文链接核对详情。"),
            )
            for row in note_rows:
                notes.append(row)
            for cell in notes[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
            for row in notes.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            notes.freeze_panes = "A2"

            workbook.save(buffer)
            date_stamp = metadata.generated_at.strftime("%Y%m%d")
            return RenderedExport(
                content=buffer.getvalue(),
                filename=safe_filename(f"AI行业情报_{date_stamp}", extension="xlsx"),
                ascii_filename=safe_filename(f"ai-intelligence-{date_stamp}", extension="xlsx"),
                media_type=self.media_type,
            )
        finally:
            workbook.close()
            buffer.close()
